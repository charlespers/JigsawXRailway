"""
BOM Exporter
Exports BOMs in multiple industry-standard formats: IPC-2581 XML, CSV, Excel
"""

import csv
import json
from typing import Dict, Any, List
from pathlib import Path
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class BOMExporter:
    """Exports BOMs in various industry-standard formats."""
    
    def __init__(self):
        pass
    
    def export_csv(
        self,
        bom: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        Export BOM to CSV format (manufacturing-friendly).
        
        Args:
            bom: BOM dictionary from OutputGenerator
            output_path: Path to save CSV file
        
        Returns:
            Path to saved file
        """
        bom_items = bom.get("items", [])
        metadata = bom.get("metadata", {})
        
        # Define CSV columns (industry-standard order)
        fieldnames = [
            "Designator",
            "Qty",
            "Manufacturer",
            "Mfr Part Number",
            "Description",
            "Category",
            "Package",
            "Footprint",
            "Value",
            "Tolerance",
            "Temperature Rating",
            "RoHS Compliant",
            "Lifecycle Status",
            "Availability Status",
            "Lead Time (days)",
            "Mounting Type",
            "Assembly Side",
            "MSL Level",
            "Datasheet URL",
            "Alternate Part Numbers",
            "Assembly Notes",
            "Unit Cost",
            "Extended Cost",
            "Notes"
        ]
        
        output_file = Path(output_path)
        if not output_file.suffix:
            output_file = output_file.with_suffix(".csv")
        
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for item in bom_items:
                # Format alternate part numbers as comma-separated string
                alt_parts = item.get("alternate_part_numbers", [])
                alt_parts_str = ", ".join(alt_parts) if isinstance(alt_parts, list) else str(alt_parts)
                
                # Format distributor part numbers
                dist_parts = item.get("distributor_part_numbers", {})
                dist_parts_str = ", ".join([f"{k}: {v}" for k, v in dist_parts.items()]) if dist_parts else ""
                
                writer.writerow({
                    "Designator": item.get("designator", ""),
                    "Qty": item.get("qty", 1),
                    "Manufacturer": item.get("manufacturer", ""),
                    "Mfr Part Number": item.get("mfr_part_number", ""),
                    "Description": item.get("description", ""),
                    "Category": item.get("category", ""),
                    "Package": item.get("package", ""),
                    "Footprint": item.get("footprint", ""),
                    "Value": item.get("value", ""),
                    "Tolerance": item.get("tolerance", ""),
                    "Temperature Rating": item.get("temperature_rating", ""),
                    "RoHS Compliant": "Yes" if item.get("rohs_compliant", True) else "No",
                    "Lifecycle Status": item.get("lifecycle_status", ""),
                    "Availability Status": item.get("availability_status", ""),
                    "Lead Time (days)": item.get("lead_time_days", ""),
                    "Mounting Type": item.get("mounting_type", ""),
                    "Assembly Side": item.get("assembly_side", ""),
                    "MSL Level": item.get("msl_level", ""),
                    "Datasheet URL": item.get("datasheet_url", ""),
                    "Alternate Part Numbers": alt_parts_str,
                    "Assembly Notes": item.get("assembly_notes", ""),
                    "Unit Cost": item.get("unit_cost", 0),
                    "Extended Cost": item.get("extended_cost", 0),
                    "Notes": item.get("notes", "")
                })
        
        return str(output_file)
    
    def export_excel(
        self,
        bom: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        Export BOM to Excel format with formatting.
        
        Args:
            bom: BOM dictionary from OutputGenerator
            output_path: Path to save Excel file
        
        Returns:
            Path to saved file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        bom_items = bom.get("items", [])
        metadata = bom.get("metadata", {})
        summary = bom.get("summary", {})
        
        output_file = Path(output_path)
        if not output_file.suffix:
            output_file = output_file.with_suffix(".xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Define columns
        columns = [
            "Designator", "Qty", "Manufacturer", "Mfr Part Number", "Description",
            "Category", "Package", "Footprint", "Value", "Tolerance",
            "Temperature Rating", "RoHS Compliant", "Lifecycle Status",
            "Availability Status", "Lead Time (days)", "Mounting Type",
            "Assembly Side", "MSL Level", "Datasheet URL", "Alternate Part Numbers",
            "Assembly Notes", "Unit Cost", "Extended Cost", "Notes"
        ]
        
        # Write header
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, item in enumerate(bom_items, start=2):
            alt_parts = item.get("alternate_part_numbers", [])
            alt_parts_str = ", ".join(alt_parts) if isinstance(alt_parts, list) else str(alt_parts)
            
            ws.cell(row=row_idx, column=1, value=item.get("designator", ""))
            ws.cell(row=row_idx, column=2, value=item.get("qty", 1))
            ws.cell(row=row_idx, column=3, value=item.get("manufacturer", ""))
            ws.cell(row=row_idx, column=4, value=item.get("mfr_part_number", ""))
            ws.cell(row=row_idx, column=5, value=item.get("description", ""))
            ws.cell(row=row_idx, column=6, value=item.get("category", ""))
            ws.cell(row=row_idx, column=7, value=item.get("package", ""))
            ws.cell(row=row_idx, column=8, value=item.get("footprint", ""))
            ws.cell(row=row_idx, column=9, value=item.get("value", ""))
            ws.cell(row=row_idx, column=10, value=item.get("tolerance", ""))
            ws.cell(row=row_idx, column=11, value=item.get("temperature_rating", ""))
            ws.cell(row=row_idx, column=12, value="Yes" if item.get("rohs_compliant", True) else "No")
            ws.cell(row=row_idx, column=13, value=item.get("lifecycle_status", ""))
            ws.cell(row=row_idx, column=14, value=item.get("availability_status", ""))
            ws.cell(row=row_idx, column=15, value=item.get("lead_time_days", ""))
            ws.cell(row=row_idx, column=16, value=item.get("mounting_type", ""))
            ws.cell(row=row_idx, column=17, value=item.get("assembly_side", ""))
            ws.cell(row=row_idx, column=18, value=item.get("msl_level", ""))
            ws.cell(row=row_idx, column=19, value=item.get("datasheet_url", ""))
            ws.cell(row=row_idx, column=20, value=alt_parts_str)
            ws.cell(row=row_idx, column=21, value=item.get("assembly_notes", ""))
            ws.cell(row=row_idx, column=22, value=item.get("unit_cost", 0))
            ws.cell(row=row_idx, column=23, value=item.get("extended_cost", 0))
            ws.cell(row=row_idx, column=24, value=item.get("notes", ""))
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(["BOM Summary"])
        summary_ws.append([])
        summary_ws.append(["Total Parts", summary.get("total_parts", 0)])
        summary_ws.append(["Total Quantity", summary.get("total_qty", 0)])
        summary_ws.append(["Total Cost", f"${summary.get('total_cost', 0):.2f}"])
        summary_ws.append([])
        summary_ws.append(["Revision", metadata.get("revision", "1.0")])
        summary_ws.append(["Revision Date", metadata.get("revision_date", "")])
        summary_ws.append(["Generated By", metadata.get("generated_by", "")])
        summary_ws.append(["Standard", metadata.get("standard", "")])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        return str(output_file)
    
    def export_ipc2581_xml(
        self,
        bom: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        Export BOM to IPC-2581 XML format (simplified version).
        
        Note: Full IPC-2581 implementation is complex. This provides a basic structure.
        
        Args:
            bom: BOM dictionary from OutputGenerator
            output_path: Path to save XML file
        
        Returns:
            Path to saved file
        """
        bom_items = bom.get("items", [])
        metadata = bom.get("metadata", {})
        summary = bom.get("summary", {})
        
        output_file = Path(output_path)
        if not output_file.suffix:
            output_file = output_file.with_suffix(".xml")
        
        xml_lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<IPC2581 xmlns="http://www.ipc2581.org/schema/IPC2581" version="B">',
            '  <LogisticHeader>',
            f'    <CompanyName>{metadata.get("generated_by", "PCB Design Agent System")}</CompanyName>',
            f'    <Revision>{metadata.get("revision", "1.0")}</Revision>',
            f'    <RevisionDate>{metadata.get("revision_date", "")}</RevisionDate>',
            '  </LogisticHeader>',
            '  <Bom>',
            '    <BomHeader>',
            f'      <BomName>PCB Design BOM</BomName>',
            f'      <BomDescription>Bill of Materials generated by PCB Design Agent System</BomDescription>',
            '    </BomHeader>',
            '    <BomItems>'
        ]
        
        for item in bom_items:
            alt_parts = item.get("alternate_part_numbers", [])
            alt_parts_str = "|".join(alt_parts) if isinstance(alt_parts, list) else str(alt_parts)
            
            xml_lines.extend([
                '      <BomItem>',
                f'        <Designator>{item.get("designator", "")}</Designator>',
                f'        <Quantity>{item.get("qty", 1)}</Quantity>',
                f'        <Manufacturer>{item.get("manufacturer", "")}</Manufacturer>',
                f'        <PartNumber>{item.get("mfr_part_number", "")}</PartNumber>',
                f'        <Description>{item.get("description", "")}</Description>',
                f'        <Package>{item.get("package", "")}</Package>',
                f'        <Footprint>{item.get("footprint", "")}</Footprint>',
                f'        <Value>{item.get("value", "")}</Value>',
                f'        <MountingType>{item.get("mounting_type", "")}</MountingType>',
                f'        <AssemblySide>{item.get("assembly_side", "")}</AssemblySide>',
                f'        <MslLevel>{item.get("msl_level", "")}</MslLevel>',
                f'        <DatasheetUrl>{item.get("datasheet_url", "")}</DatasheetUrl>',
                f'        <AlternatePartNumbers>{alt_parts_str}</AlternatePartNumbers>',
                f'        <AssemblyNotes>{item.get("assembly_notes", "")}</AssemblyNotes>',
                '      </BomItem>'
            ])
        
        xml_lines.extend([
            '    </BomItems>',
            '    <BomSummary>',
            f'      <TotalParts>{summary.get("total_parts", 0)}</TotalParts>',
            f'      <TotalQuantity>{summary.get("total_qty", 0)}</TotalQuantity>',
            f'      <TotalCost>{summary.get("total_cost", 0)}</TotalCost>',
            '    </BomSummary>',
            '  </Bom>',
            '</IPC2581>'
        ])
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(xml_lines))
        
        return str(output_file)
    
    def export_json(
        self,
        bom: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        Export BOM to JSON format (for programmatic use).
        
        Args:
            bom: BOM dictionary from OutputGenerator
            output_path: Path to save JSON file
        
        Returns:
            Path to saved file
        """
        output_file = Path(output_path)
        if not output_file.suffix:
            output_file = output_file.with_suffix(".json")
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(bom, f, indent=2, ensure_ascii=False)
        
        return str(output_file)

